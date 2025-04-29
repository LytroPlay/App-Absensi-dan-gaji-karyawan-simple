import sys
import os
from datetime import datetime, timedelta
import calendar
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                            QLabel, QPushButton, QTabWidget, QTableWidget, QTableWidgetItem, 
                            QLineEdit, QFormLayout, QMessageBox, QComboBox, QDateEdit, QDialog,
                            QDialogButtonBox, QGroupBox, QGridLayout, QFileDialog)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QIcon
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import locale

# Set locale untuk format mata uang (titik sebagai pemisah ribuan)
locale.setlocale(locale.LC_ALL, 'id_ID.UTF-8')

class EmployeeDialog(QDialog):
    def __init__(self, parent=None, employee_data=None):
        super().__init__(parent)
        self.setWindowTitle("Data Karyawan")
        self.setMinimumWidth(300)
        
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()
        
        self.name_input = QLineEdit()
        self.base_salary_input = QLineEdit()
        self.overtime_rate_input = QLineEdit()
        
        form_layout.addRow("Nama:", self.name_input)
        form_layout.addRow("Gaji Pokok (Rp/Jam):", self.base_salary_input)
        form_layout.addRow("Lembur (Rp/Jam):", self.overtime_rate_input)
        
        layout.addLayout(form_layout)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        
        layout.addWidget(buttons)
        
        self.setLayout(layout)
        
        # Jika ada data karyawan, isi form
        if employee_data:
            self.name_input.setText(employee_data[0])
            self.base_salary_input.setText(str(employee_data[1]))
            self.overtime_rate_input.setText(str(employee_data[2]))
    
    def get_employee_data(self):
        try:
            name = self.name_input.text()
            base_salary = int(self.base_salary_input.text())
            overtime_rate = int(self.overtime_rate_input.text())
            return [name, base_salary, overtime_rate]
        except ValueError:
            QMessageBox.warning(self, "Error", "Gaji dan tarif lembur harus berupa angka!")
            return None

class AttendanceDialog(QDialog):
    def __init__(self, parent=None, employees=None, date=None):
        super().__init__(parent)
        self.setWindowTitle("Input Absensi")
        self.setMinimumWidth(400)
        
        self.employees = employees
        self.employee_attendance = {}
        
        main_layout = QVBoxLayout()
        
        # Tanggal absensi
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Tanggal:"))
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate() if date is None else date)
        date_layout.addWidget(self.date_edit)
        main_layout.addLayout(date_layout)
        
        # Group box untuk absensi
        group_box = QGroupBox("Absensi Karyawan")
        grid_layout = QGridLayout()
        
        # Header
        grid_layout.addWidget(QLabel("Nama"), 0, 0)
        grid_layout.addWidget(QLabel("Status"), 0, 1)
        grid_layout.addWidget(QLabel("Jam Kerja"), 0, 2)
        grid_layout.addWidget(QLabel("Jam Lembur"), 0, 3)
        
        # Row untuk tiap karyawan
        for i, employee in enumerate(self.employees):
            grid_layout.addWidget(QLabel(employee[0]), i+1, 0)
            
            status_combo = QComboBox()
            status_combo.addItems(["Masuk", "Tidak Masuk"])
            grid_layout.addWidget(status_combo, i+1, 1)
            
            work_hours = QLineEdit("8")  # Default 8 jam
            grid_layout.addWidget(work_hours, i+1, 2)
            
            overtime_hours = QLineEdit("0")  # Default 0 jam lembur
            grid_layout.addWidget(overtime_hours, i+1, 3)
            
            self.employee_attendance[employee[0]] = {
                'status_combo': status_combo,
                'work_hours': work_hours,
                'overtime_hours': overtime_hours
            }
            
            # Connect signal untuk mengaktifkan/menonaktifkan input jam
            status_combo.currentTextChanged.connect(
                lambda text, wh=work_hours, oh=overtime_hours: self.toggle_hour_inputs(text, wh, oh))
        
        group_box.setLayout(grid_layout)
        main_layout.addWidget(group_box)
        
        # Tombol OK/Cancel
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        main_layout.addWidget(buttons)
        
        self.setLayout(main_layout)
    
    def toggle_hour_inputs(self, status, work_hours, overtime_hours):
        enabled = status == "Masuk"
        work_hours.setEnabled(enabled)
        overtime_hours.setEnabled(enabled)
        if not enabled:
            work_hours.setText("0")
            overtime_hours.setText("0")
    
    def get_attendance_data(self):
        date = self.date_edit.date().toString("yyyy-MM-dd")
        attendance_data = []
        
        for employee_name, inputs in self.employee_attendance.items():
            status = inputs['status_combo'].currentText()
            try:
                work_hours = int(inputs['work_hours'].text()) if status == "Masuk" else 0
                overtime_hours = int(inputs['overtime_hours'].text()) if status == "Masuk" else 0
                attendance_data.append([date, employee_name, status, work_hours, overtime_hours])
            except ValueError:
                QMessageBox.warning(self, "Error", f"Jam kerja untuk {employee_name} harus berupa angka!")
                return None
        
        return attendance_data

class LaundryPayrollApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.check_and_create_files()
        self.load_employee_data()
        
    def initUI(self):
        self.setWindowTitle("Sistem Absensi & Gaji Pilot Laundry")
        self.setGeometry(100, 100, 900, 600)
        
        # Widget utama dan layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout()
        main_widget.setLayout(main_layout)
        
        # Tab widget untuk navigasi
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        # Tab Karyawan
        self.employee_tab = QWidget()
        employee_layout = QVBoxLayout()
        self.employee_tab.setLayout(employee_layout)
        
        # Tabel karyawan
        self.employee_table = QTableWidget()
        self.employee_table.setColumnCount(3)
        self.employee_table.setHorizontalHeaderLabels(["Nama", "Gaji Pokok (Rp/Jam)", "Lembur (Rp/Jam)"])
        self.employee_table.horizontalHeader().setStretchLastSection(True)
        employee_layout.addWidget(self.employee_table)
        
        # Tombol untuk tab karyawan
        employee_buttons = QHBoxLayout()
        self.add_employee_btn = QPushButton("Tambah Karyawan")
        self.add_employee_btn.clicked.connect(self.add_employee)
        self.edit_employee_btn = QPushButton("Edit Karyawan")
        self.edit_employee_btn.clicked.connect(self.edit_employee)
        self.delete_employee_btn = QPushButton("Hapus Karyawan")
        self.delete_employee_btn.clicked.connect(self.delete_employee)
        
        employee_buttons.addWidget(self.add_employee_btn)
        employee_buttons.addWidget(self.edit_employee_btn)
        employee_buttons.addWidget(self.delete_employee_btn)
        employee_layout.addLayout(employee_buttons)
        
        # Tab Absensi
        self.attendance_tab = QWidget()
        attendance_layout = QVBoxLayout()
        self.attendance_tab.setLayout(attendance_layout)
        
        # Widget untuk filter tanggal
        date_filter = QHBoxLayout()
        date_filter.addWidget(QLabel("Tanggal:"))
        self.attendance_date = QDateEdit()
        self.attendance_date.setCalendarPopup(True)
        self.attendance_date.setDate(QDate.currentDate())
        self.attendance_date.dateChanged.connect(self.load_attendance_data)
        date_filter.addWidget(self.attendance_date)
        attendance_layout.addLayout(date_filter)
        
        # Tabel absensi
        self.attendance_table = QTableWidget()
        self.attendance_table.setColumnCount(5)
        self.attendance_table.setHorizontalHeaderLabels(["Tanggal", "Nama", "Status", "Jam Kerja", "Jam Lembur"])
        self.attendance_table.horizontalHeader().setStretchLastSection(True)
        attendance_layout.addWidget(self.attendance_table)
        
        # Tombol untuk tab absensi
        attendance_buttons = QHBoxLayout()
        self.add_attendance_btn = QPushButton("Input Absensi")
        self.add_attendance_btn.clicked.connect(self.add_attendance)
        self.delete_attendance_btn = QPushButton("Hapus Absensi")
        self.delete_attendance_btn.clicked.connect(self.delete_attendance)
        
        attendance_buttons.addWidget(self.add_attendance_btn)
        attendance_buttons.addWidget(self.delete_attendance_btn)
        attendance_layout.addLayout(attendance_buttons)
        
        # Tab Laporan Gaji
        self.salary_tab = QWidget()
        salary_layout = QVBoxLayout()
        self.salary_tab.setLayout(salary_layout)
        
        # Widget untuk rentang tanggal
        date_range = QHBoxLayout()
        date_range.addWidget(QLabel("Dari:"))
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addDays(-30))  # Default 30 hari sebelumnya
        date_range.addWidget(self.from_date)
        
        date_range.addWidget(QLabel("Sampai:"))
        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        date_range.addWidget(self.to_date)
        
        self.generate_report_btn = QPushButton("Hitung Gaji")
        self.generate_report_btn.clicked.connect(self.generate_salary_report)
        date_range.addWidget(self.generate_report_btn)
        
        self.export_excel_btn = QPushButton("Export Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        date_range.addWidget(self.export_excel_btn)
        
        salary_layout.addLayout(date_range)
        
        # Tabel laporan gaji
        self.salary_table = QTableWidget()
        self.salary_table.setColumnCount(7)
        self.salary_table.setHorizontalHeaderLabels([
            "Nama", "Total Hari", "Total Jam", "Total Lembur", 
            "Gaji Pokok", "Gaji Lembur", "Total Gaji"
        ])
        self.salary_table.horizontalHeader().setStretchLastSection(True)
        salary_layout.addWidget(self.salary_table)
        
        # Tambahkan semua tab ke tab widget
        self.tabs.addTab(self.employee_tab, "Data Karyawan")
        self.tabs.addTab(self.attendance_tab, "Absensi")
        self.tabs.addTab(self.salary_tab, "Laporan Gaji")
        
        # Load data awal
        self.show()
    
    def check_and_create_files(self):
        # Cek dan buat file excel jika belum ada
        if not os.path.exists("data"):
            os.makedirs("data")
        
        # File karyawan
        if not os.path.exists("data/employees.xlsx"):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employees"
            ws.append(["Nama", "Gaji Pokok", "Lembur"])
            wb.save("data/employees.xlsx")
        
        # File absensi
        if not os.path.exists("data/attendance.xlsx"):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Tanggal", "Nama", "Status", "Jam Kerja", "Jam Lembur"])
            wb.save("data/attendance.xlsx")
    
    def load_employee_data(self):
        try:
            wb = openpyxl.load_workbook("data/employees.xlsx")
            ws = wb["Employees"]
            
            # Clear tabel terlebih dahulu
            self.employee_table.setRowCount(0)
            
            self.employees = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Skip baris kosong
                    self.employees.append(row)
                    row_position = self.employee_table.rowCount()
                    self.employee_table.insertRow(row_position)
                    
                    self.employee_table.setItem(row_position, 0, QTableWidgetItem(str(row[0])))
                    
                    # Format angka dengan pemisah ribuan (titik)
                    salary_item = QTableWidgetItem(f"{row[1]:,}".replace(",", "."))
                    salary_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.employee_table.setItem(row_position, 1, salary_item)
                    
                    overtime_item = QTableWidgetItem(f"{row[2]:,}".replace(",", "."))
                    overtime_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.employee_table.setItem(row_position, 2, overtime_item)
            
            # Resize kolom agar sesuai dengan konten
            self.employee_table.resizeColumnsToContents()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data karyawan: {e}")
    
    def load_attendance_data(self):
        try:
            selected_date = self.attendance_date.date().toString("yyyy-MM-dd")
            
            wb = openpyxl.load_workbook("data/attendance.xlsx")
            ws = wb["Attendance"]
            
            # Clear tabel terlebih dahulu
            self.attendance_table.setRowCount(0)
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == selected_date:
                    row_position = self.attendance_table.rowCount()
                    self.attendance_table.insertRow(row_position)
                    
                    for col, value in enumerate(row):
                        item = QTableWidgetItem(str(value))
                        self.attendance_table.setItem(row_position, col, item)
            
            # Resize kolom agar sesuai dengan konten
            self.attendance_table.resizeColumnsToContents()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data absensi: {e}")
    
    def add_employee(self):
        dialog = EmployeeDialog(self)
        if dialog.exec_():
            employee_data = dialog.get_employee_data()
            if employee_data:
                try:
                    wb = openpyxl.load_workbook("data/employees.xlsx")
                    ws = wb["Employees"]
                    ws.append(employee_data)
                    wb.save("data/employees.xlsx")
                    
                    QMessageBox.information(self, "Sukses", "Data karyawan berhasil ditambahkan!")
                    self.load_employee_data()
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Gagal menambahkan data karyawan: {e}")
    
    def edit_employee(self):
        selected_row = self.employee_table.currentRow()
        if selected_row >= 0:
            employee_name = self.employee_table.item(selected_row, 0).text()
            
            # Cari data karyawan yang akan diedit
            employee_data = None
            for emp in self.employees:
                if emp[0] == employee_name:
                    employee_data = emp
                    break
            
            if employee_data:
                dialog = EmployeeDialog(self, employee_data)
                if dialog.exec_():
                    new_data = dialog.get_employee_data()
                    if new_data:
                        try:
                            wb = openpyxl.load_workbook("data/employees.xlsx")
                            ws = wb["Employees"]
                            
                            # Update data di excel
                            for row in ws.iter_rows(min_row=2):
                                if row[0].value == employee_name:
                                    row[0].value = new_data[0]
                                    row[1].value = new_data[1]
                                    row[2].value = new_data[2]
                                    break
                            
                            wb.save("data/employees.xlsx")
                            
                            # Update juga data absensi jika nama berubah
                            if new_data[0] != employee_name:
                                self.update_attendance_employee_name(employee_name, new_data[0])
                            
                            QMessageBox.information(self, "Sukses", "Data karyawan berhasil diupdate!")
                            self.load_employee_data()
                        except Exception as e:
                            QMessageBox.critical(self, "Error", f"Gagal mengupdate data karyawan: {e}")
        else:
            QMessageBox.warning(self, "Peringatan", "Pilih karyawan yang akan diedit terlebih dahulu!")
    
    def update_attendance_employee_name(self, old_name, new_name):
        try:
            wb = openpyxl.load_workbook("data/attendance.xlsx")
            ws = wb["Attendance"]
            
            for row in ws.iter_rows(min_row=2):
                if row[1].value == old_name:
                    row[1].value = new_name
            
            wb.save("data/attendance.xlsx")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal mengupdate nama karyawan di absensi: {e}")
    
    def delete_employee(self):
        selected_row = self.employee_table.currentRow()
        if selected_row >= 0:
            employee_name = self.employee_table.item(selected_row, 0).text()
            
            reply = QMessageBox.question(self, "Konfirmasi", 
                                        f"Yakin ingin menghapus karyawan {employee_name}?",
                                        QMessageBox.Yes | QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                try:
                    wb = openpyxl.load_workbook("data/employees.xlsx")
                    ws = wb["Employees"]
                    
                    # Cari dan hapus baris yang sesuai
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if row[0].value == employee_name:
                            ws.delete_rows(row_idx)
                            break
                    
                    wb.save("data/employees.xlsx")
                    
                    QMessageBox.information(self, "Sukses", "Data karyawan berhasil dihapus!")
                    self.load_employee_data()
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Gagal menghapus data karyawan: {e}")
        else:
            QMessageBox.warning(self, "Peringatan", "Pilih karyawan yang akan dihapus terlebih dahulu!")
    
    def add_attendance(self):
        if not self.employees:
            QMessageBox.warning(self, "Peringatan", "Tidak ada data karyawan! Tambahkan karyawan terlebih dahulu.")
            return
        
        selected_date = self.attendance_date.date()
        
        # Cek apakah sudah ada data absensi di tanggal tersebut
        wb = openpyxl.load_workbook("data/attendance.xlsx")
        ws = wb["Attendance"]
        date_str = selected_date.toString("yyyy-MM-dd")
        
        date_exists = False
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == date_str:
                date_exists = True
                break
        
        if date_exists:
            reply = QMessageBox.question(self, "Konfirmasi", 
                                        f"Data absensi tanggal {date_str} sudah ada. Timpa data?",
                                        QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.No:
                return
            else:
                # Hapus data lama
                rows_to_delete = []
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value == date_str:
                        rows_to_delete.append(row_idx)
                
                # Hapus dari bawah ke atas agar indeks tidak berubah
                for row_idx in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(row_idx)
                wb.save("data/attendance.xlsx")
        
        dialog = AttendanceDialog(self, self.employees, selected_date)
        if dialog.exec_():
            attendance_data = dialog.get_attendance_data()
            if attendance_data:
                try:
                    wb = openpyxl.load_workbook("data/attendance.xlsx")
                    ws = wb["Attendance"]
                    
                    # Tambahkan data absensi baru
                    for att in attendance_data:
                        ws.append(att)
                    
                    wb.save("data/attendance.xlsx")
                    QMessageBox.information(self, "Sukses", "Data absensi berhasil disimpan!")
                    self.load_attendance_data()
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Gagal menyimpan data absensi: {e}")
    
    def delete_attendance(self):
        selected_date = self.attendance_date.date().toString("yyyy-MM-dd")
        
        reply = QMessageBox.question(self, "Konfirmasi", 
                                    f"Yakin ingin menghapus semua data absensi tanggal {selected_date}?",
                                    QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                wb = openpyxl.load_workbook("data/attendance.xlsx")
                ws = wb["Attendance"]
                
                # Cari dan hapus baris yang sesuai
                rows_to_delete = []
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value == selected_date:
                        rows_to_delete.append(row_idx)
                
                # Hapus dari bawah ke atas agar indeks tidak berubah
                for row_idx in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(row_idx)
                
                wb.save("data/attendance.xlsx")
                
                QMessageBox.information(self, "Sukses", "Data absensi berhasil dihapus!")
                self.load_attendance_data()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal menghapus data absensi: {e}")
    
    def generate_salary_report(self):
        try:
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            
            # Load data karyawan
            wb_emp = openpyxl.load_workbook("data/employees.xlsx")
            ws_emp = wb_emp["Employees"]
            
            employees = {}
            for row in ws_emp.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    employees[row[0]] = {
                        'base_salary': row[1],
                        'overtime_rate': row[2],
                        'work_days': 0,
                        'work_hours': 0,
                        'overtime_hours': 0
                    }
            
            # Load data absensi
            wb_att = openpyxl.load_workbook("data/attendance.xlsx")
            ws_att = wb_att["Attendance"]
            
            for row in ws_att.iter_rows(min_row=2, values_only=True):
                date_str, name, status, work_hours, overtime_hours = row
                
                # Filter berdasarkan rentang tanggal
                if from_date <= date_str <= to_date and name in employees and status == "Masuk":
                    employees[name]['work_days'] += 1
                    employees[name]['work_hours'] += work_hours
                    employees[name]['overtime_hours'] += overtime_hours
            
            # Hitung gaji
            salary_data = []
            for name, data in employees.items():
                base_salary_total = data['base_salary'] * data['work_hours']
                overtime_total = data['overtime_rate'] * data['overtime_hours']
                total_salary = base_salary_total + overtime_total
                
                salary_data.append([
                    name,
                    data['work_days'],
                    data['work_hours'],
                    data['overtime_hours'],
                    base_salary_total,
                    overtime_total,
                    total_salary
                ])
            
            # Tampilkan di tabel
            self.salary_table.setRowCount(0)
            for row_data in salary_data:
                row_position = self.salary_table.rowCount()
                self.salary_table.insertRow(row_position)
                
                # Nama
                self.salary_table.setItem(row_position, 0, QTableWidgetItem(row_data[0]))
                
                # Hari & Jam
                for col in range(1, 4):
                    item = QTableWidgetItem(str(row_data[col]))
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.salary_table.setItem(row_position, col, item)
                
                # Gaji (dengan format mata uang)
                for col in range(4, 7):
                    item = QTableWidgetItem(f"Rp {row_data[col]:,}".replace(",", "."))
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.salary_table.setItem(row_position, col, item)
            
            # Resize kolom
            self.salary_table.resizeColumnsToContents()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal membuat laporan gaji: {e}")
    
    def export_to_excel(self):
        try:
            # Dapatkan path untuk menyimpan file
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(
                self, "Simpan Laporan Excel", "", "Excel Files (*.xlsx)", options=options
            )
            
            if not file_name:
                return
            
            if not file_name.endswith('.xlsx'):
                file_name += '.xlsx'
            
            # Buat workbook baru
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laporan Gaji"
            
            # Header periode
            from_date = self.from_date.date().toString("dd/MM/yyyy")
            to_date = self.to_date.date().toString("dd/MM/yyyy")
            
            ws['A1'] = "LAPORAN GAJI KARYAWAN PILOT LAUNDRY"
            ws.merge_cells('A1:G1')
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws['A2'] = f"Periode: {from_date} s/d {to_date}"
            ws.merge_cells('A2:G2')
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Header tabel
            headers = ["Nama", "Total Hari", "Total Jam", "Total Lembur", 
                       "Gaji Pokok", "Gaji Lembur", "Total Gaji"]
            
            for col, header in enumerate(headers):
                cell = ws.cell(row=4, column=col+1)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                # Tambahkan border
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
            
            # Isi data
            row_num = 5
            for i in range(self.salary_table.rowCount()):
                for j in range(self.salary_table.columnCount()):
                    cell = ws.cell(row=row_num, column=j+1)
                    
                    # Dapatkan nilai dari tabel
                    value = self.salary_table.item(i, j).text()
                    
                    # Jika kolom gaji, hapus "Rp " dan konversi ke angka
                    if j >= 4:
                        if "Rp " in value:
                            value = value.replace("Rp ", "").replace(".", "")
                        cell.value = int(value)
                        cell.number_format = '#,##0'  # Format angka dengan pemisah ribuan
                    else:
                        try:
                            cell.value = int(value)
                        except ValueError:
                            cell.value = value
                    
                    # Alignment sesuai tipe data
                    if j == 0:  # Nama
                        cell.alignment = Alignment(horizontal='left')
                    else:  # Angka
                        cell.alignment = Alignment(horizontal='right')
                    
                    # Tambahkan border
                    cell.border = thin_border
                
                row_num += 1
            
            # Auto-fit kolom - dengan cara yang lebih aman
            for i, col in enumerate(ws.columns, 1):
                max_length = 0
                column = openpyxl.utils.get_column_letter(i)
                
                for cell in col:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        # Abaikan sel yang tidak bisa diakses
                        pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Simpan workbook
            wb.save(file_name)
            
            QMessageBox.information(self, "Sukses", f"Laporan berhasil disimpan di {file_name}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal mengekspor laporan: {e}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = LaundryPayrollApp()
    sys.exit(app.exec_())